"""
Admin Bulk Operations Router — endpoints for importing and exporting Inventory, Menu Items, and Customers.
"""
from __future__ import annotations

import io

from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from app.dependencies import DBSession, RequireAdmin
from app.schemas.bulk_operations import BulkImportSummary, ExportFormatEnum
from app.services.bulk_export_service import export_customers, export_inventory, export_menu_items
from app.services.bulk_import_service import import_customers, import_inventory, import_menu_items

router = APIRouter(prefix="/api/admin/bulk", tags=["admin-bulk-operations"])

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


async def _read_and_validate_file(file: UploadFile) -> bytes:
    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="File size exceeds the 5MB limit."
        )
    return file_bytes


# ---------------------------------------------------------
# Inventory
# ---------------------------------------------------------

@router.post("/inventory/import", response_model=BulkImportSummary)
async def import_inventory_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    file: UploadFile = File(...),
):
    """Import inventory items from CSV/Excel and auto-create related Menu Items and Categories."""
    file_bytes = await _read_and_validate_file(file)
    try:
        return await import_inventory(db, current_user.outlet_id, file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/inventory/export")
async def export_inventory_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    format: ExportFormatEnum = Query(ExportFormatEnum.EXCEL),
):
    """Export all inventory items to CSV, Excel, or PDF."""
    file_bytes, filename, content_type = await export_inventory(db, current_user.outlet_id, format)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------
# Menu Items
# ---------------------------------------------------------

@router.post("/menu-items/import", response_model=BulkImportSummary)
async def import_menu_items_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    file: UploadFile = File(...),
):
    """Import menu items from CSV/Excel directly (without inventory linkage)."""
    file_bytes = await _read_and_validate_file(file)
    try:
        return await import_menu_items(db, current_user.outlet_id, file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/menu-items/export")
async def export_menu_items_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    format: ExportFormatEnum = Query(ExportFormatEnum.EXCEL),
):
    """Export all menu items to CSV, Excel, or PDF."""
    file_bytes, filename, content_type = await export_menu_items(db, current_user.outlet_id, format)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------
# Customers
# ---------------------------------------------------------

@router.post("/customers/import", response_model=BulkImportSummary)
async def import_customers_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    file: UploadFile = File(...),
):
    """Import customers from CSV/Excel. Optionally creates legacy orders for historical spend."""
    file_bytes = await _read_and_validate_file(file)
    try:
        return await import_customers(db, current_user.outlet_id, file_bytes, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/customers/export")
async def export_customers_endpoint(
    current_user: RequireAdmin,
    db: DBSession,
    format: ExportFormatEnum = Query(ExportFormatEnum.EXCEL),
):
    """Export all customers with aggregated stats to CSV, Excel, or PDF."""
    file_bytes, filename, content_type = await export_customers(db, current_user.outlet_id, format)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------
# Templates
# ---------------------------------------------------------

@router.get("/templates/{entity}")
async def download_template(
    current_user: RequireAdmin,
    entity: str,
):
    """Download a blank template with the correct headers and formulas for import."""
    import pandas as pd
    
    if entity == "inventory":
        cols = ["Name", "Barcode", "Unit", "Category", "Initial Qty", "Sorted Qty", "Total Billed", "Cost Per Unit", "MRP Margin Pct", "Retail Margin Pct", "Wholesale Margin Pct", "MRP", "Retail Price", "Wholesale Price", "Margin Type", "Supplier", "Expiry Date", "Tax Category", "Tax Rate", "Reorder Threshold", "Shelf Life Alert Hrs", "Current Stock"]
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Protection
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Template"
        
        # Write headers
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
        # Add formula to "Cost Per Unit" (Column H = 8)
        # E = Initial Qty, F = Sorted Qty, G = Total Billed
        for row_idx in range(2, 501):
            formula = f'=IF(F{row_idx}>0, G{row_idx}/F{row_idx}, IF(E{row_idx}>0, G{row_idx}/E{row_idx}, ""))'
            cell = ws.cell(row=row_idx, column=8, value=formula)
            # Optional: lock the formula cell so user doesn't overwrite it
            cell.protection = Protection(locked=True)
            
            # Add formulas for MRP (L=12), Retail Price (M=13), Wholesale Price (N=14)
            # using Margin Type (O=15), MRP Margin Pct (I=9), Retail Margin Pct (J=10), Wholesale Margin Pct (K=11)
            ws.cell(row=row_idx, column=12, value=f'=IF(ISNUMBER(I{row_idx}), IF(O{row_idx}="MARGIN", H{row_idx}/(1-I{row_idx}/100), H{row_idx}+(H{row_idx}*I{row_idx}/100)), "")')
            ws.cell(row=row_idx, column=13, value=f'=IF(ISNUMBER(J{row_idx}), IF(O{row_idx}="MARGIN", H{row_idx}/(1-J{row_idx}/100), H{row_idx}+(H{row_idx}*J{row_idx}/100)), "")')
            ws.cell(row=row_idx, column=14, value=f'=IF(ISNUMBER(K{row_idx}), IF(O{row_idx}="MARGIN", H{row_idx}/(1-K{row_idx}/100), H{row_idx}+(H{row_idx}*K{row_idx}/100)), "")')
            
        # Protect worksheet but allow inserting rows and selecting cells
        ws.protection.sheet = True
        ws.protection.formatCells = False
        ws.protection.insertRows = False
        # Unlock all cells except the ones we explicitly locked
        for row in ws.iter_rows(min_row=2, max_row=500, min_col=1, max_col=len(cols)):
            for cell in row:
                if cell.column not in [8, 12, 13, 14]:  # If not a formula column
                    cell.protection = Protection(locked=False)
        
        # Adjust column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
            
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventory_template.xlsx"}
        )

    elif entity == "menu-items":
        cols = ["Name", "Category", "Price", "Barcode", "Description", "MRP", "Wholesale Price", "Evening Price", "Offer Price", "Offer Label", "Tax Category", "Tax Rate", "Pricing Mode", "Unit Label", "Is Available"]
    elif entity == "customers":
        cols = ["Phone", "Name", "Loyalty Points", "Historical Spend"]
    else:
        raise HTTPException(status_code=404, detail="Unknown entity. Use 'inventory', 'menu-items', or 'customers'.")
        
    df = pd.DataFrame(columns=cols)
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={entity}_template.csv"}
    )
